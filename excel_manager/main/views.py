import os
import io
import shutil
import datetime
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, decorators
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.template import TemplateDoesNotExist
from django.conf import settings
from django.http import HttpResponse, Http404, HttpResponseForbidden
from django.shortcuts import redirect, render, get_object_or_404
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
import pandas as pd

from .models import ExcelFile
from .forms import ExcelUploadForm, OrganizationCreationForm

def index(request):
    """Главная страница с формой входа и кнопкой перехода к бета-версии."""
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # Если пользователь – администратор, перенаправляем в админпанель
            if user.is_superuser:
                return redirect('admin_dashboard')
            else:
                # Если организация – перенаправляем на её дашборд
                return redirect('organization_detail', org_id=user.id)
    else:
        form = AuthenticationForm()
    return render(request, 'index.html', {'form': form})

def beta(request):
    """Страница бета-версии с меню."""
    return render(request, 'beta.html')

def beta_page(request, page_name):
    """
    Отображает страницу из каталога project_beta по имени.
    Если шаблон project_beta/<page_name>.html не найден,
    возвращается страница-заглушка.
    """
    template_path = f'project_beta/{page_name}.html'
    try:
        # Проверяем, существует ли такой шаблон
        get_template(template_path)
        # Если шаблон найден, рендерим его
        return render(request, template_path)
    except TemplateDoesNotExist:
        # Либо можно вернуть 404, либо отрендерить заглушку
        context = {'page_name': page_name}
        return render(request, 'stub_page.html', context)

@decorators.login_required
def download_excel(request):
    """Скачивание Excel‑файла с сохранением форматирования."""
    user = request.user
    excel_files = user.excel_files.all()
    excel_file = excel_files.first() if excel_files.exists() else None

    if not excel_file:
        messages.error(request, "Здесь пока что нет информации")
        return redirect('dashboard')

    try:
        wb = load_workbook(excel_file.file.path)
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response = HttpResponse(
            virtual_workbook.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{excel_file.title}.xlsx"'
        return response
    except Exception as e:
        messages.error(request, f"Ошибка при скачивании файла: {e}")
        return redirect('dashboard')

@decorators.login_required
def admin_dashboard(request):
    """
    Панель администратора:
    – Загрузка Excel‑файла (без ввода названия)
    – Создание учётных записей организаций (только логин, пароль и наименование)
    – Список организаций с возможностью просмотра, удаления и переключения доступа к Excel‑файлу
    – Список загруженного Excel‑файла с кнопкой удаления
    """
    # Обработка загрузки Excel‑файла
    if request.method == 'POST' and 'upload_excel' in request.POST:
        upload_form = ExcelUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            file_name = upload_form.cleaned_data['file'].name
            file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', file_name)

            # Удаляем старый файл, если существует
            if os.path.exists(file_path):
                os.remove(file_path)

            # Если файл существует в базе, обновляем его
            existing = ExcelFile.objects.filter(title=file_name).first()
            if existing:
                existing.file.delete()  # Удаляем старый файл из базы данных
                existing.file = upload_form.cleaned_data['file']  # Присваиваем новый файл
                existing.save()
            else:
                # Если файла нет — создаём новый
                instance = upload_form.save(commit=False)
                instance.title = file_name
                instance.save()

            messages.success(request, "Excel‑файл успешно загружен.")
            return redirect('admin_dashboard')  # Убедитесь, что перенаправление происходит после добавления сообщения
    else:
        upload_form = ExcelUploadForm()

    # Обработка создания организации
    if request.method == 'POST' and 'create_org' in request.POST:
        org_form = OrganizationCreationForm(request.POST)
        if org_form.is_valid():
            # Проверка на существование логина
            username = org_form.cleaned_data['username']
            if User.objects.filter(username=username).exists():
                messages.error(request, "Такой логин организации уже занят!")
            else:
                org_form.save()
                messages.success(request, "Организация успешно создана.")
                return redirect('admin_dashboard')  # Убедитесь, что перенаправление происходит после добавления сообщения
    else:
        org_form = OrganizationCreationForm()

    # Получение первого Excel файла и списка организаций
    excel_file = ExcelFile.objects.first()
    organizations = User.objects.filter(is_superuser=False)

    context = {
        'upload_form': upload_form,
        'org_form': org_form,
        'excel_file': excel_file,
        'organizations': organizations,
    }
    return render(request, 'admin_dashboard.html', context)

@decorators.login_required
def delete_excel_file(request, file_id):
    excel_file = get_object_or_404(ExcelFile, id=file_id)
    excel_file.delete()
    messages.success(request, "Excel‑файл удален.")
    return redirect('admin_dashboard')

def is_row_numeric(row):
    """
    Возвращает True, если все заполненные ячейки строки (из объектов ячеек) можно привести к числу.
    """
    for cell in row:
        val = cell.value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            continue
        try:
            float(val)
        except (ValueError, TypeError):
            return False
    return True

def is_numeric_list(values):
    """
    Возвращает True, если все непустые значения из списка values можно привести к числу.
    """
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        try:
            float(v)
        except (ValueError, TypeError):
            return False
    return True

def ws_to_html(ws):
    """
    Преобразует лист Excel в HTML-таблицу с учетом следующих требований:
      - Даты форматируются как ДД.ММ.ГГГГ (без времени).
      - Первый столбец не выводится.
      - При поиске строки-заголовка пропускаются строки, которые либо полностью пустые, 
        либо все заполненные ячейки можно привести к числу (например, строка "1	2	3	4	5	6	7	8	9	11	12	13	14").
      - Строки данных, где все отображаемые ячейки пусты или все значения числовые, не выводятся.
      - Столбцы с пустым заголовком, с "=TODAY()" или "Отметка о выполнении" не выводятся.
      - Если в строке данных встречается значение "Срок" или "Факт" в одном из разрешённых столбцов,
        такая строка не включается.
      - В ячейках столбца "Примечание", если значение равно "Закрыто", фон ячейки окрашивается в светло-зелёный.
      - Для столбцов с заголовками "Срок" и "Факт" добавляется класс "nowrap", чтобы их содержимое не переносилось.
    """
    # Поиск строки заголовка: перебираем строки с начала, пропуская полностью пустые и полностью числовые строки.
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            continue
        if is_row_numeric(row):
            continue
        header_row_idx = row[0].row
        break

    if header_row_idx is None:
        return "<p>Нет данных</p>"

    # Получаем строку заголовков
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
    
    allowed_columns = []  # Индексы столбцов для отображения
    headers = []          # Тексты заголовков для этих столбцов
    for i, cell in enumerate(header_cells):
        # Пропускаем первый столбец
        if i == 0:
            continue
        cell_value = cell.value if cell.value is not None else ""
        cell_text = str(cell_value).strip()
        # Пропускаем столбцы с пустым заголовком, "=TODAY()" или "Отметка о выполнении"
        if cell_text == "" or cell_text == "=TODAY()" or cell_text == "Отметка о выполнении":
            continue
        allowed_columns.append(i)
        headers.append(cell_text)

    html = f'<table class="table table-bordered table-striped" id="table_{ws.title}">'
    html += '<thead><tr>'
    for header in headers:
        if header in ("Срок", "Факт"):
            html += f'<th class="nowrap">{header}</th>'
        else:
            html += f'<th>{header}</th>'
    html += '</tr></thead><tbody>'
    
    # Строки данных начинаются со следующей строки после заголовков
    data_start = header_row_idx + 1
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row):
        # Собираем значения для разрешённых столбцов
        row_values = []
        for col_idx in allowed_columns:
            cell = row[col_idx] if col_idx < len(row) else None
            val = cell.value if cell is not None else None
            row_values.append(val)
        # Пропускаем строку, если все отображаемые ячейки пусты
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue
        # Пропускаем строку, если все непустые значения можно привести к числу
        if is_numeric_list(row_values):
            continue
        # Если в строке встречается "Срок" или "Факт", пропускаем строку
        if any(str(v).strip() in ("Срок", "Факт") for v in row_values if v is not None):
            continue

        html += '<tr>'
        for j, col_idx in enumerate(allowed_columns):
            cell = row[col_idx]
            value = cell.value if cell is not None else ""
            if value is None:
                value = ""
            else:
                if isinstance(value, (datetime.datetime, datetime.date)):
                    value = value.strftime("%d.%m.%Y")
            if headers[j] == "Примечание" and str(value).strip() == "Закрыто":
                html += f'<td style="background-color: lightgreen;">{value}</td>'
            elif headers[j] in ("Срок", "Факт"):
                html += f'<td class="nowrap">{value}</td>'
            else:
                html += f'<td>{value}</td>'
        html += '</tr>'
    html += '</tbody></table>'
    return html

@login_required
def organization_detail(request, org_id):
    org = get_object_or_404(User, id=org_id)
    if not request.user.is_superuser and request.user.id != org.id:
        return HttpResponse("Доступ запрещен.", status=403)

    excel_file = ExcelFile.objects.first()
    if not excel_file or org not in excel_file.allowed_organizations.all():
        #messages.error(request, "Нет доступа к Excel‑файлу.")
        return render(request, 'organization_detail.html', {'org': org, 'file_url': None, 'tables': None})

    orig_file_path = excel_file.file.path
    target_dir = os.path.join(settings.MEDIA_ROOT, org.username)
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    target_filename = os.path.basename(orig_file_path)
    target_file_path = os.path.join(target_dir, target_filename)

    shutil.copy2(orig_file_path, target_file_path)

    try:
        wb = load_workbook(target_file_path)
        org_name = str(org.first_name).strip().lower()
        found = False  # Флаг: найден ли нужный столбец
        tables = {}

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            target_col_index = None

            # Поиск столбца "Наименование строительной подрядной организации"
            for cell in ws[1]:
                if cell.value == "Наименование строительной подрядной организации":
                    if isinstance(cell.column, int):
                        target_col_index = cell.column
                    else:
                        target_col_index = column_index_from_string(cell.column)
                    found = True
                    break

            # Если столбец найден, удаляем лишние строки
            if target_col_index:
                for row in range(ws.max_row, 4, -1):
                    cell_val = ws.cell(row=row, column=target_col_index).value
                    if cell_val is None or str(cell_val).strip().lower() != org_name:
                        ws.delete_rows(row)
                # Добавляем HTML-таблицу в контекст
                tables[sheet] = ws_to_html(ws)

        if not found:
            messages.error(request, "Не найден нужный столбец на ни одном листе.")
            file_url = None
        else:
            wb.save(target_file_path)
            file_url = os.path.join(settings.MEDIA_URL, org.username, target_filename)

    except Exception as e:
        messages.error(request, f"Ошибка обработки файла: {e}")
        file_url = None
        tables = None

    context = {
        'org': org,
        'file_url': file_url,
        'tables': tables,
    }
    return render(request, 'organization_detail.html', context)

@login_required
def organization_download_excel(request, org_id):
    """
    Скачивание отфильтрованного Excel‑файла для конкретной организации.

    Функция копирует оригинальный Excel‑файл в папку media/{org.username},
    затем проверяет все листы на наличие столбца "Наименование строительной подрядной организации".
    В найденных листах удаляются строки (с 5-й и ниже), если значение в этом столбце
    не совпадает с org.first_name, и сохраняет файл.

    Затем файл отдается пользователю для скачивания.
    """
    org = get_object_or_404(User, id=org_id)
    if not request.user.is_superuser and request.user.id != org.id:
        return HttpResponse("Доступ запрещен.", status=403)

    excel_file = ExcelFile.objects.first()
    if not excel_file or org not in excel_file.allowed_organizations.all():
        messages.error(request, "Нет доступа к Excel‑файлу.")
        return redirect('organization_detail', org_id=org.id)

    orig_file_path = excel_file.file.path
    target_dir = os.path.join(settings.MEDIA_ROOT, org.username)
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    target_filename = os.path.basename(orig_file_path)
    target_file_path = os.path.join(target_dir, target_filename)

    shutil.copy2(orig_file_path, target_file_path)

    try:
        wb = load_workbook(target_file_path)

        org_name = str(org.first_name).strip().lower()
        found = False  # Флаг, нашли ли мы нужный столбец хотя бы в одном листе

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            target_col_index = None

            # Поиск нужного столбца
            for cell in ws[1]:
                if cell.value == "Наименование строительной подрядной организации":
                    if isinstance(cell.column, int):
                        target_col_index = cell.column
                    else:
                        target_col_index = column_index_from_string(cell.column)
                    found = True
                    break

            # Если нашли, удаляем ненужные строки
            if target_col_index:
                for row in range(ws.max_row, 4, -1):
                    cell_val = ws.cell(row=row, column=target_col_index).value
                    if cell_val is None or str(cell_val).strip().lower() != org_name:
                        ws.delete_rows(row)

        if not found:
            messages.error(request, "Не найден нужный столбец на ни одном листе.")
            return redirect('organization_detail', org_id=org.id)

        wb.save(target_file_path)

        with open(target_file_path, 'rb') as f:
            file_data = f.read()
        response = HttpResponse(
            file_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{org.first_name}_filtered.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f"Ошибка при скачивании файла: {e}")
        return redirect('organization_detail', org_id=org.id)

@decorators.login_required
def delete_organization(request, org_id):
    org = get_object_or_404(User, id=org_id)
    if request.method == 'POST':
        org.delete()
        messages.success(request, "Организация удалена.")
    return redirect('admin_dashboard')

@decorators.login_required
def toggle_excel_access(request, org_id):
    org = get_object_or_404(User, id=org_id)
    excel_file = ExcelFile.objects.first()  # так как файл один
    if excel_file:
        if org in excel_file.allowed_organizations.all():
            excel_file.allowed_organizations.remove(org)
            messages.info(request, f"Доступ Excel‑файла для {org.username} отключен.")
        else:
            excel_file.allowed_organizations.add(org)
            messages.success(request, f"Доступ Excel‑файла для {org.username} включен.")
    else:
        messages.error(request, "Нет загруженного Excel‑файла.")
    return redirect('admin_dashboard')

def user_logout(request):
    logout(request)
    return redirect('index')
